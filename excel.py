
from openpyxl import Workbook, load_workbook


# Write Excel


wb=Workbook(write_only=True)
ws= wb.create_sheet('vikas',index=1)
ws1= wb.create_sheet('anand',index=0)

for i in range(100):
    ws.append(['a','b','c','d','e'])

wb.save("myexcel.xlsx")

wb= load_workbook(filename='myexcel.xlsx',read_only=True)
ws = wb['vikas']

for row in ws.iter_rows():
    l1=[]
    for cell in row:
        print(cell.coordinate,"==")
        l1.append(cell.value)

    l1.append(l1)


wb = load_workbook(filename='myexcel.xlsx')
ws = wb['vikas']

for row in ws.iter_rows():
    for cell in row:
        if cell.coordinate == 'A99':
            ws[cell.coordinate] = "Vikas"
            print(ws[cell.coordinate])

wb.save("myexcel.xlsx")